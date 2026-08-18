"""
Motore deterministico SONAR (componente C4, cap. 7 del documento di progetto).

Legge sonar_registry.xlsx, calcola FIT / TIMING / PONTE, applica le regole
anti-ICP e instrada ogni azienda in coda A/B/C/D/Quarantena/Fuori.

Non genera testo (motivazione, One Pager, messaggi): quella parte e' agentica
e la fa Claude leggendo i segnali, non questo script (principio della
proposta di architettura, sez. 1: "non tutto deve essere un agente LLM").

Uso:
    python sonar_engine.py score                 # ricalcola tutte le aziende
    python sonar_engine.py score --id AZ-000001   # ricalcola una sola azienda
    python sonar_engine.py queue --cap 25         # esporta la coda della settimana
    python sonar_engine.py check                  # valida coerenza dati, non scrive nulla
"""
import argparse
import csv
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
WORKBOOK_PATH = DATA_DIR / "sonar_registry.xlsx"


# ---------------------------------------------------------------------------
# Lettura generica fogli come liste di dict
# ---------------------------------------------------------------------------

def sheet_to_dicts(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = rows[0]
    out = []
    for r in rows[1:]:
        if all(v is None for v in r):
            continue
        out.append(dict(zip(headers, r)))
    return out


def col_index(ws, header_name):
    for cell in ws[1]:
        if cell.value == header_name:
            return cell.column
    raise KeyError(f"Colonna '{header_name}' non trovata nel foglio {ws.title}")


def truthy(v):
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    return str(v).strip().lower() in ("si", "true", "1", "x", "yes")


def to_number(v, default=0):
    if v is None or v == "":
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def to_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

def load_config(wb):
    fit_rows = sheet_to_dicts(wb["Config_Pesi_FIT"])
    timing_rows = sheet_to_dicts(wb["Config_Pesi_TIMING"])
    soglie_rows = sheet_to_dicts(wb["Config_Soglie_Instradamento"])
    timing_by_code = {r["codice_segnale"]: r for r in timing_rows}
    return {
        "fit": fit_rows,          # [{famiglia, segnale, campo_flag_in_Aziende, punti}, ...]
        "timing_by_code": timing_by_code,
        "soglie": {r["coda"]: r for r in soglie_rows},
    }


# ---------------------------------------------------------------------------
# Anti-ICP (par. 2.2 / 2.3.3 vincolo / 7.3 regola 1)
# ---------------------------------------------------------------------------

ANTI_ICP_FIELDS = [
    "anti_icp_fatturato_o_struttura",
    "anti_icp_funzione_interna_strutturata",
    "anti_icp_contratto_manutenzione_attivo",
    "anti_icp_cliente_atman_attivo",
    "anti_icp_conflitto_interesse",
]
OPPOSIZIONE_FIELD = "anti_icp_opposizione_registrata"


def anti_icp_status(azienda):
    flags = sum(1 for f in ANTI_ICP_FIELDS if truthy(azienda.get(f)))
    opposizione = truthy(azienda.get(OPPOSIZIONE_FIELD))
    if opposizione:
        return flags, "esclusione_opposizione"
    if flags >= 2:
        return flags, "esclusione_anti_icp"
    if flags == 1:
        return flags, "quarantena"
    return flags, "ok"


# ---------------------------------------------------------------------------
# FIT (par. 7.2, somma di 9 criteri booleani = max 100)
# ---------------------------------------------------------------------------

def compute_fit(azienda, config):
    total = 0
    for row in config["fit"]:
        campo = row["campo_flag_in_Aziende"]
        if truthy(azienda.get(campo)):
            total += int(row["punti"])
    return min(total, 100)


# ---------------------------------------------------------------------------
# TIMING (par. 7.2-7.3, somma con decadimento, limitata a 100)
# Regola di decadimento: pieno entro la finestra di validita', meta' fino al
# doppio della finestra, zero oltre. I segnali permanenti non decadono.
# I segnali "fino a scadenza" (S6) valgono pieno fino a data_scadenza inclusa.
# ---------------------------------------------------------------------------

def decayed_points(segnale, cfg_row, oggi):
    punti_base = cfg_row["punti"]
    permanente = str(cfg_row.get("permanente", "")).strip().lower()

    if permanente == "si":
        return punti_base

    if permanente == "fino a scadenza":
        scadenza = to_date(segnale.get("data_scadenza"))
        if scadenza is None:
            return 0  # dato mancante: non calcolabile, va corretto a mano
        return punti_base if oggi <= scadenza else 0

    finestra = cfg_row.get("finestra_validita_giorni")
    rilevazione = to_date(segnale.get("data_rilevazione"))
    if finestra is None or rilevazione is None:
        return 0
    eta_giorni = (oggi - rilevazione).days
    if eta_giorni < 0:
        return 0
    if eta_giorni <= finestra:
        return punti_base
    if eta_giorni <= 2 * finestra:
        return punti_base / 2
    return 0


def compute_timing(azienda_id, segnali_rows, config, oggi):
    total = 0.0
    dettaglio = []
    for s in segnali_rows:
        if s.get("azienda_id") != azienda_id:
            continue
        if str(s.get("contribuisce_a", "")).strip().upper() != "TIMING":
            continue
        cfg_row = config["timing_by_code"].get(s.get("codice_segnale"))
        if cfg_row is None:
            continue
        pts = decayed_points(s, cfg_row, oggi)
        total += pts
        if pts > 0:
            dettaglio.append((s.get("codice_segnale"), pts, s.get("data_rilevazione"), s.get("fonte_url")))
    return min(total, 100), dettaglio


# ---------------------------------------------------------------------------
# PONTE (cap. 6, valore massimo fra i collegamenti rilevati)
# ---------------------------------------------------------------------------

def compute_ponte(azienda_id, ponti_rows):
    forze = [to_number(p.get("forza_ponte_0_100")) for p in ponti_rows if p.get("azienda_id") == azienda_id]
    return max(forze) if forze else 0


# ---------------------------------------------------------------------------
# Instradamento (par. 7.4)
# ---------------------------------------------------------------------------

def route(fit, timing, ponte, anti_icp_stato):
    if anti_icp_stato == "esclusione_opposizione":
        return "Fuori (opposizione)"
    if anti_icp_stato == "esclusione_anti_icp":
        return "Fuori (anti-ICP)"
    if anti_icp_stato == "quarantena":
        return "Quarantena"
    if fit < 30:
        return "Fuori (FIT<30)"
    if ponte >= 50 and fit >= 60:
        return "A"
    if fit >= 60 and timing >= 60:
        return "B"
    if fit >= 60:
        return "C"
    if 30 <= fit < 60 and ponte >= 60:
        return "D"
    return "Fuori (residuale)"


# ---------------------------------------------------------------------------
# Comando: score
# ---------------------------------------------------------------------------

def cmd_score(args):
    wb = load_workbook(WORKBOOK_PATH)
    config = load_config(wb)
    ws_az = wb["Aziende"]
    segnali_rows = sheet_to_dicts(wb["Segnali"])
    ponti_rows = sheet_to_dicts(wb["Ponti"])
    oggi = date.today()

    idx = {
        "fit": col_index(ws_az, "fit_score"),
        "timing": col_index(ws_az, "timing_score"),
        "ponte": col_index(ws_az, "ponte_score"),
        "coda": col_index(ws_az, "coda"),
        "flagcount": col_index(ws_az, "anti_icp_flag_count"),
        "data_calcolo": col_index(ws_az, "data_calcolo"),
        "azienda_id": col_index(ws_az, "azienda_id"),
    }

    n = 0
    rows = list(ws_az.iter_rows(min_row=2))
    for row in rows:
        azienda_id_cell = row[idx["azienda_id"] - 1]
        azienda_id = azienda_id_cell.value
        if azienda_id is None:
            continue
        if args.id and azienda_id != args.id:
            continue

        azienda = {ws_az.cell(row=1, column=c.column).value: c.value for c in row}

        flags, stato = anti_icp_status(azienda)
        fit = compute_fit(azienda, config)
        timing, _ = compute_timing(azienda_id, segnali_rows, config, oggi)
        ponte = compute_ponte(azienda_id, ponti_rows)
        coda = route(fit, timing, ponte, stato)

        row[idx["fit"] - 1].value = fit
        row[idx["timing"] - 1].value = round(timing, 1)
        row[idx["ponte"] - 1].value = ponte
        row[idx["coda"] - 1].value = coda
        row[idx["flagcount"] - 1].value = flags
        row[idx["data_calcolo"] - 1].value = oggi.isoformat()
        n += 1

    wb.save(WORKBOOK_PATH)
    print(f"Ricalcolate {n} aziende. Coda/FIT/TIMING/PONTE aggiornati in {WORKBOOK_PATH.name}.")
    if n == 0:
        print("Nessuna azienda trovata (o nessun azienda_id corrisponde a --id). "
              "Il foglio 'Aziende' e' ancora vuoto?")


# ---------------------------------------------------------------------------
# Comando: queue - esporta la coda della settimana rispettando il tetto (par. 8.5)
# ---------------------------------------------------------------------------

def cmd_queue(args):
    wb = load_workbook(WORKBOOK_PATH)
    ws_az = wb["Aziende"]
    aziende = sheet_to_dicts(ws_az)

    code_a = sorted([a for a in aziende if a.get("coda") == "A"],
                     key=lambda a: (-(a.get("ponte_score") or 0), -(a.get("fit_score") or 0)))
    code_b = sorted([a for a in aziende if a.get("coda") == "B"],
                     key=lambda a: (-(a.get("fit_score") or 0), -(a.get("timing_score") or 0)))
    code_d = sorted([a for a in aziende if a.get("coda") == "D"],
                     key=lambda a: -(a.get("ponte_score") or 0))

    selected = (code_a + code_b + code_d)[: args.cap]

    out_path = DATA_DIR / f"coda_settimana_{date.today().isoformat()}.csv"
    fields = ["azienda_id", "ragione_sociale", "coda", "fit_score", "timing_score",
              "ponte_score", "prodotto_frontend", "owner_commerciale", "motivazione"]
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for a in selected:
            writer.writerow(a)

    print(f"Coda della settimana: {len(selected)} aziende (tetto {args.cap}) -> {out_path}")
    print(f"  Coda A (presentazione): {len(code_a)} disponibili, {min(len(code_a), args.cap)} incluse")
    print(f"  Coda B (contatto diretto): {len(code_b)} disponibili")
    print(f"  Coda D (condizionata a ponte forte): {len(code_d)} disponibili")
    if not selected:
        print("Nessuna azienda in coda A/B/D. Esegui prima 'sonar_engine.py score'.")


# ---------------------------------------------------------------------------
# Comando: check - validazione di coerenza, nessuna scrittura
# ---------------------------------------------------------------------------

def cmd_check(args):
    wb = load_workbook(WORKBOOK_PATH)
    aziende = sheet_to_dicts(wb["Aziende"])
    segnali = sheet_to_dicts(wb["Segnali"])
    ponti = sheet_to_dicts(wb["Ponti"])

    problems = []
    ids = set()
    for a in aziende:
        aid = a.get("azienda_id")
        if not aid:
            problems.append("Riga in Aziende senza azienda_id")
            continue
        if aid in ids:
            problems.append(f"azienda_id duplicato: {aid}")
        ids.add(aid)
        if not a.get("partita_iva"):
            problems.append(f"{aid}: partita_iva mancante (chiave di deduplica primaria, par. 4.4)")

    for s in segnali:
        if s.get("azienda_id") not in ids:
            problems.append(f"Segnale {s.get('segnale_id')} punta a azienda_id inesistente: {s.get('azienda_id')}")
        if s.get("contribuisce_a") == "TIMING" and not s.get("codice_segnale"):
            problems.append(f"Segnale {s.get('segnale_id')}: manca codice_segnale, non contribuira' al TIMING")

    for p in ponti:
        if p.get("azienda_id") not in ids:
            problems.append(f"Ponte {p.get('ponte_id')} punta a azienda_id inesistente: {p.get('azienda_id')}")

    riserva = [a for a in aziende if truthy(a.get("riserva_confronto"))]
    print(f"Aziende: {len(aziende)} | Segnali: {len(segnali)} | Ponti: {len(ponti)} | Riserva di confronto: {len(riserva)}")
    if aziende:
        quota = len(riserva) / len(aziende)
        if quota < 0.10 and len(aziende) >= 20:
            problems.append(f"Riserva di confronto al {quota:.0%}, sotto il 10-15% richiesto dal par. 7.5")

    if problems:
        print(f"\n{len(problems)} problema/i trovato/i:")
        for p in problems:
            print(f"  - {p}")
        sys.exit(1)
    print("Nessun problema di coerenza rilevato.")


def main():
    parser = argparse.ArgumentParser(description="Motore SONAR (C4 - punteggi e instradamento)")
    sub = parser.add_subparsers(dest="command", required=True)

    p_score = sub.add_parser("score", help="ricalcola FIT/TIMING/PONTE/coda")
    p_score.add_argument("--id", default=None, help="ricalcola solo questo azienda_id")
    p_score.set_defaults(func=cmd_score)

    p_queue = sub.add_parser("queue", help="esporta la coda di lavoro della settimana")
    p_queue.add_argument("--cap", type=int, default=25, help="tetto settimanale (par. 3.3: coda corta, max 25)")
    p_queue.set_defaults(func=cmd_queue)

    p_check = sub.add_parser("check", help="valida la coerenza dei dati senza scrivere")
    p_check.set_defaults(func=cmd_check)

    args = parser.parse_args()
    if not WORKBOOK_PATH.exists():
        print(f"Workbook non trovato: {WORKBOOK_PATH}. Esegui prima build_workbook.py")
        sys.exit(1)
    args.func(args)


if __name__ == "__main__":
    main()
