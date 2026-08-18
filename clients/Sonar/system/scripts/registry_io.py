"""
Utility di lettura/scrittura sicura per sonar_registry.xlsx.

Pensata per essere importata da snippet Python brevi lanciati da Claude Code
via Bash durante il lavoro sulla pipeline (C1-C6), cosi' da non dover mai
scrivere righe posizionali a mano (fonte di errori, vedi test del motore).

Esempi d'uso (da riga di comando, dentro system/scripts/):

    python -c "
from registry_io import open_wb, save_wb, add_row, next_id
wb = open_wb()
aid = next_id(wb, 'Aziende', 'azienda_id', 'AZ')
add_row(wb, 'Aziende', azienda_id=aid, ragione_sociale='Rossi Srl',
        partita_iva='IT01234567890', ateco_primario='25.11',
        fit_ateco_target=True, fit_fatturato_in_range=True, ...)
save_wb(wb)
"
"""
from pathlib import Path
from openpyxl import load_workbook

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
WORKBOOK_PATH = DATA_DIR / "sonar_registry.xlsx"


def open_wb():
    return load_workbook(WORKBOOK_PATH)


def save_wb(wb):
    wb.save(WORKBOOK_PATH)


def headers_of(wb, sheet):
    ws = wb[sheet]
    return [c.value for c in ws[1]]


def next_id(wb, sheet, id_field, prefix, width=6):
    ws = wb[sheet]
    hdrs = headers_of(wb, sheet)
    idx = hdrs.index(id_field)
    max_n = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[idx]
        if val and str(val).startswith(prefix + "-"):
            try:
                n = int(str(val).split("-")[-1])
                max_n = max(max_n, n)
            except ValueError:
                pass
    return f"{prefix}-{max_n + 1:0{width}d}"


def add_row(wb, sheet, **fields):
    """Aggiunge una riga usando SOLO nomi di campo (mai posizione).
    I campi non passati restano vuoti. Solleva errore se un nome di campo
    non esiste nell'intestazione — cosi' un typo si nota subito."""
    ws = wb[sheet]
    hdrs = headers_of(wb, sheet)
    unknown = set(fields) - set(hdrs)
    if unknown:
        raise KeyError(f"Campi non presenti nel foglio '{sheet}': {sorted(unknown)}")
    row = [fields.get(h) for h in hdrs]
    ws.append(row)
    return row


def update_row(wb, sheet, id_field, id_value, **fields):
    """Aggiorna i campi indicati sulla riga identificata da id_field=id_value.
    Non tocca le colonne non passate. Torna True se ha trovato la riga."""
    ws = wb[sheet]
    hdrs = headers_of(wb, sheet)
    id_idx = hdrs.index(id_field)
    unknown = set(fields) - set(hdrs)
    if unknown:
        raise KeyError(f"Campi non presenti nel foglio '{sheet}': {sorted(unknown)}")
    for row in ws.iter_rows(min_row=2):
        if row[id_idx].value == id_value:
            for name, value in fields.items():
                row[hdrs.index(name)].value = value
            return True
    return False


def find_rows(wb, sheet, **where):
    """Ritorna una lista di dict per le righe che soddisfano tutte le
    condizioni where (uguaglianza semplice). Utile per controlli rapidi
    prima di inserire (es. deduplica su partita_iva, par. 4.4)."""
    ws = wb[sheet]
    hdrs = headers_of(wb, sheet)
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(hdrs, row))
        if all(d.get(k) == v for k, v in where.items()):
            out.append(d)
    return out
