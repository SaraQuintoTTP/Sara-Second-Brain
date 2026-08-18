"""
Inizializza sonar_registry.xlsx: lo storage strutturato del sistema SONAR.
Esegui una sola volta (per Sprint 0). Se il file esiste gia', si ferma senza
sovrascrivere, a meno di --force (che ricrea SOLO i fogli Config_*, mai i dati).

Uso:
    python build_workbook.py            # crea il workbook se non esiste
    python build_workbook.py --force    # rigenera i fogli Config_* (pesi/soglie/ICP)
                                         # lasciando intatti Aziende/Contatti/Ponti/Segnali/Esiti
"""
import sys
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
WORKBOOK_PATH = DATA_DIR / "sonar_registry.xlsx"

HEADER_FILL = PatternFill(start_color="1A2D40", end_color="1A2D40", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CONFIG_FILL = PatternFill(start_color="F5F2EE", end_color="F5F2EE", fill_type="solid")


def write_sheet(wb, name, headers, rows=None, col_widths=None):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    if rows:
        for row in rows:
            ws.append(row)
    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return ws


# ---------------------------------------------------------------------------
# Fogli dati operativi (vuoti alla creazione, popolati durante il lavoro)
# ---------------------------------------------------------------------------

AZIENDE_HEADERS = [
    "azienda_id", "ragione_sociale", "partita_iva", "forma_giuridica",
    "sede_legale", "sedi_operative",
    "ateco_primario", "ateco_secondari", "settore_atman", "vertical",
    "descrizione_attivita", "area_dominante", "prodotto_frontend",
    "fatturato_ultimo_bilancio", "anno_bilancio", "addetti", "trend_triennale_positivo",
    "dominio", "piattaforma_cms", "ecommerce_transazionale", "data_ultima_modifica_sito",
    "certificato_ssl", "gestore_consenso", "tracciamento_presente", "dichiarazione_accessibilita",
    "recensioni_n", "recensioni_media", "ultima_recensione", "canali_social_attivi",
    "ultima_pubblicazione_social", "visibilita_ricerca",
    # --- flag booleani usati dal motore per FIT (par. 7.2) ---
    "fit_ateco_target", "fit_fatturato_in_range", "fit_addetti_in_range",
    "fit_sede_priorita1", "fit_trend_positivo", "fit_asset_digitale_proprietario",
    "fit_ecommerce_attivo", "fit_stack_compatibile", "fit_decisore_raggiungibile",
    # --- flag booleani usati dal motore per anti-ICP (par. 2.2) ---
    "anti_icp_fatturato_o_struttura", "anti_icp_funzione_interna_strutturata",
    "anti_icp_contratto_manutenzione_attivo", "anti_icp_cliente_atman_attivo",
    "anti_icp_conflitto_interesse", "anti_icp_opposizione_registrata",
    # --- output calcolati dal motore (sonar_engine.py) — NON compilare a mano ---
    "fit_score", "timing_score", "ponte_score", "coda", "anti_icp_flag_count",
    "motivazione", "data_calcolo",
    # --- relazione e governance ---
    "origine_record", "stato", "owner_commerciale",
    "data_raccolta", "fonte_puntuale", "base_giuridica", "scadenza_conservazione",
    "riserva_confronto", "note",
]
AZIENDE_WIDTHS = [10, 28, 14, 14, 22, 22, 12, 16, 16, 14, 30, 16, 18,
                  16, 10, 8, 12, 14, 14, 16, 16, 12, 14, 14, 16,
                  10, 12, 14, 16, 16, 14,
                  10, 12, 10, 10, 10, 12, 10, 10, 10,
                  12, 12, 12, 12, 12, 12,
                  8, 10, 10, 8, 14, 40, 12,
                  16, 10, 16, 12, 16, 14, 16, 10, 30]

CONTATTI_HEADERS = ["contatto_id", "azienda_id", "nome", "ruolo", "livello_L0_L4",
                     "recapito", "canale", "data_verifica", "esito_verifica", "note"]

PONTI_HEADERS = ["ponte_id", "azienda_id", "tipo_ponte", "forza_ponte_0_100",
                  "sorgente", "soggetto_origine", "persona_atman_riferimento",
                  "data_rilevazione", "stato_utilizzo", "esito_presentazione", "note"]

SEGNALI_HEADERS = ["segnale_id", "azienda_id", "famiglia", "codice_segnale",
                    "descrizione", "contribuisce_a", "valore_punti",
                    "data_rilevazione", "fonte_url", "finestra_validita_giorni",
                    "permanente", "data_scadenza"]

ESITI_HEADERS = ["esito_id", "azienda_id", "data", "coda_provenienza",
                  "origine_contatto", "tipo_ponte_usato", "cluster_messaggio",
                  "canale", "variante_messaggio", "n_tentativi", "esito",
                  "motivo_rifiuto", "valore_contratto_eur",
                  "fit_snapshot", "timing_snapshot", "ponte_snapshot",
                  "riserva_confronto"]

# ---------------------------------------------------------------------------
# Fogli di configurazione — trascritti 1:1 dal documento di progetto v3.2
# ---------------------------------------------------------------------------

CONFIG_ICP = [
    ["Geografia", "Veneto. P1: Padova, Venezia, Treviso, Vicenza. P2: Verona, Rovigo, Belluno. P3: Friuli-Venezia Giulia", "Registro Imprese", "Deciso"],
    ["Dimensione", "Fatturato 500.000-20.000.000 euro, addetti 10-50", "Bilanci depositati e banche dati camerali", "Deciso"],
    ["Settore", "Manifattura e servizi B2B; produzione B2B; studi professionali; medicale e odontoiatrico; food e commercio locale evoluto; pet", "Codice ATECO con verifica sul sito", "Deciso"],
    ["Governance", "Societa' di capitali o studio professionale strutturato, proprieta' identificabile, decisore raggiungibile senza filtro corporate", "Visura, sito, profili professionali pubblici", "Proposto"],
    ["Maturita' digitale", "Dispone gia' di almeno un asset digitale proprietario", "Analisi tecnica del dominio", "Proposto"],
    ["Bisogno atteso", "Sviluppo/rifacimento sito ed e-commerce; automazioni/integrazioni gestionali; sostituzione fornitore non strutturato; digitalizzazione processi/canali di vendita; gestione social", "Segnali (cap. 7)", "Deciso"],
]

CONFIG_ANTI_ICP = [
    ["Fatturato sotto 500 mila euro o assenza di struttura organizzativa", "Non sostenibile un contratto ricorsivo; il progetto una tantum satura il team a bassa marginalita'"],
    ["Impresa con funzione marketing o IT interna strutturata e procedure formali di fornitura", "Ciclo di vendita lungo con procedura comparativa; territorio presidiato dai concorrenti full-service"],
    ["Contratto di manutenzione attivo e recente con altro fornitore", "Costo di sostituzione elevato. Esclusione temporanea, con data di rilettura fissata alla scadenza presunta del contratto"],
    ["Cliente ATMAN attivo", "Appartiene al flusso di sviluppo del portafoglio, che ha owner e messaggi diversi"],
    ["Conflitto di interesse dichiarato con un cliente in portafoglio", "Vincolo di relazione. Esclusione, salvo autorizzazione scritta della direzione"],
    ["Opposizione registrata o richiesta di non essere ricontattato", "Vincolo di legge. Implementato come blocco tecnico, non come annotazione"],
]

CONFIG_PESI_FIT = [
    ["S1. Struttura", "ATECO in uno dei settori target", "fit_ateco_target", 15],
    ["S1. Struttura", "Fatturato nella fascia 0,5-20 M euro", "fit_fatturato_in_range", 15],
    ["S1. Struttura", "Addetti fra 10 e 50", "fit_addetti_in_range", 10],
    ["S1. Struttura", "Sede in provincia di priorita' 1", "fit_sede_priorita1", 10],
    ["S1. Struttura", "Andamento del fatturato positivo negli ultimi tre bilanci", "fit_trend_positivo", 10],
    ["S2. Maturita' digitale", "Dispone di un asset digitale proprietario", "fit_asset_digitale_proprietario", 10],
    ["S2. Maturita' digitale", "E-commerce transazionale attivo", "fit_ecommerce_attivo", 15],
    ["S2. Maturita' digitale", "Stack tecnico compatibile con le competenze ATMAN", "fit_stack_compatibile", 10],
    ["S1. Struttura", "Decisore identificabile e raggiungibile senza filtro corporate", "fit_decisore_raggiungibile", 5],
]

CONFIG_PESI_TIMING = [
    ["S3. Movimento", "Annuncio di assunzione in area marketing, e-commerce o IT", "S3_ASSUNZIONE", 90, 25, "No"],
    ["S3. Movimento", "Accesso a bando, contributo o incentivo con vincolo di spesa", "S3_BANDO", 180, 20, "No"],
    ["S3. Movimento", "Cambio di referente o di ruolo apicale", "S3_REFERENTE", 180, 20, "No"],
    ["S3. Movimento", "Partecipazione a fiera di settore", "S3_FIERA", 120, 10, "No"],
    ["S3. Movimento", "Nuova sede, nuova linea di prodotto, cambio di identita' visiva", "S3_NUOVA_SEDE", 180, 10, "No"],
    ["S5. Fornitore", "Fornitore identificato risultato cessato o non piu' operativo", "S5_FORNITORE_CESSATO", None, 25, "Si"],
    ["S5. Fornitore", "Nessuna traccia di manutenzione da oltre 24 mesi", "S5_NO_MANUTENZIONE_24M", None, 15, "Si"],
    ["S6. Normativa", "Adempimento con scadenza applicabile all'azienda", "S6_ADEMPIMENTO", None, 15, "Fino a scadenza"],
    ["S4. Reputazione", "Presidio abbandonato: nessuna pubblicazione o risposta alle recensioni", "S4_PRESIDIO_ABBANDONATO", None, 5, "Si"],
]

CONFIG_SOGLIE = [
    ["A", "Ponte >= 50 e FIT >= 60", "Richiesta di presentazione tramite la persona di riferimento. Nessun contatto a freddo", "3-6"],
    ["B", "FIT >= 60 e TIMING >= 60, Ponte < 50", "Studio dell'azienda, One Pager, richiesta di appuntamento", "7-14"],
    ["C", "FIT >= 60 e TIMING < 60, Ponte < 50", "Coltivazione: contenuti, inviti a eventi. Rilettura a 90 giorni", "0 (no assegnazione)"],
    ["D", "FIT 30-59", "Lavorabile solo in presenza di un Ponte >= 60", "residuale"],
    ["Fuori", "FIT < 30, oppure anti-ICP (>=2 regole), oppure opposizione registrata", "Esclusione", "0"],
]

CONFIG_PONTE_TIPOLOGIE = [
    ["Persona transitata da un cliente ATMAN all'azienda in lista, o viceversa", "Profili professionali pubblici, storico referenti in Ruko", 100],
    ["Segnalatore del programma partner che conosce direttamente il decisore", "Elenco segnalatori attivi", 90],
    ["Socio, amministratore o membro dell'organo di controllo in comune", "Visura camerale", 80],
    ["Professionista, consulente o fornitore di ATMAN con rapporto diretto sull'azienda in lista", "Rubrica rete professionale", 70],
    ["Fornitore o cliente in comune, stessa filiera o stesso distretto", "Sito, comunicazioni pubbliche", 60],
    ["Stesso consulente/commercialista/associazione condivisi con un cliente ATMAN, senza rapporto diretto con ATMAN", "Conoscenza diretta del team, elenchi associativi", 50],
    ["Stessa fiera, stesso consorzio, stesso evento", "Elenchi espositori e partecipanti", 40],
    ["Stesso settore con caso studio ATMAN spendibile, senza persona in comune", "Portfolio ATMAN", 20],
    ["Nessun collegamento rilevato", "-", 0],
]

CONFIG_SEGNALE_PRODOTTO = [
    ["Fornitore cessato o non piu' operativo; nessuna traccia di manutenzione", "Tecnologia", "Audit Light (o contatto diretto se necessita' dichiarata)", "Restyling o nuovo sito", "Contratto di continuita'"],
    ["E-commerce attivo senza manutenzione, senza conformita' o su piattaforma obsoleta", "Tecnologia", "Audit Completo Tecnologia", "Rifacimento e messa in conformita'", "Contratto full con monitoraggio"],
    ["Adempimento normativo con scadenza applicabile", "Tecnologia", "Audit Light con focus conformita'", "Intervento mirato", "Contratto light"],
    ["Annuncio di assunzione in area marketing o e-commerce", "Marketing e comunicazione", "Audit Light Marketing", "Strategia e impostazione dei canali", "Gestione social e consulenza continuativa"],
    ["Presidio abbandonato: nessuna pubblicazione, nessuna risposta alle recensioni", "Marketing e comunicazione", "Audit Light Marketing", "Piano editoriale e impostazione", "Gestione social"],
    ["Partecipazione a fiera di settore o lancio di una nuova linea", "Marketing e comunicazione", "Audit Light (Audit Completo se il progetto e' rilevante)", "Materiali e canali per il lancio", "Presidio continuativo"],
    ["Crescita di fatturato e organico con processi ancora manuali", "Processi aziendali", "Audit Completo Processi", "Integrazione fra software o automazione", "Manutenzione evolutiva"],
    ["Nuova sede, riorganizzazione, cambio di gestionale", "Processi aziendali", "Audit Light Processi", "Integrazione software", "Contratto di assistenza"],
    ["Bando, contributo o incentivo ottenuto con vincolo di spesa", "Secondo l'oggetto del bando", "Audit Light area corrispondente (Completo se progetto rilevante)", "Progetto finanziato", "Continuita' sul progetto realizzato"],
    ["Necessita' dichiarata esplicitamente dal decisore", "Secondo la necessita'", "Contatto diretto", "Preventivo diretto", "Da proporre alla consegna"],
]

README_TEXT = [
    ["SONAR — registro operativo"],
    [""],
    ["Questo workbook e' il datastore del sistema SONAR per ATMAN (Sprint 0-1)."],
    ["Sostituisce il 'foglio strutturato' previsto dal documento di progetto v3.2, cap. 11."],
    [""],
    ["Fogli dati (si popolano lavorando):"],
    ["  Aziende   - un record per azienda, schema Allegato B"],
    ["  Contatti  - un record per contatto (C2)"],
    ["  Ponti     - un record per collegamento relazionale rilevato (C3)"],
    ["  Segnali   - un record per segnale rilevato, anche multipli per azienda (C4)"],
    ["  Esiti     - un record per esito registrato, snapshot congelato (C6)"],
    [""],
    ["Fogli di configurazione (NON toccare senza motivo — sono presi 1:1 dal documento; modificarli e' una decisione di retrospettiva, par. 7.5):"],
    ["  Config_ICP, Config_AntiICP, Config_Pesi_FIT, Config_Pesi_TIMING,"],
    ["  Config_Soglie_Instradamento, Config_Ponte_Tipologie, Config_Segnale_Prodotto"],
    [""],
    ["Il calcolo di FIT / TIMING / PONTE e l'instradamento in coda si fanno con:"],
    ["  python system/scripts/sonar_engine.py score"],
    ["Non modificare mai a mano le colonne fit_score / timing_score / ponte_score / coda:"],
    ["  sono sovrascritte ad ogni esecuzione del motore."],
]


def build(force=False):
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    if WORKBOOK_PATH.exists() and not force:
        print(f"Il file esiste gia': {WORKBOOK_PATH}")
        print("Uso --force per rigenerare SOLO i fogli Config_* (i dati restano intatti).")
        return

    if WORKBOOK_PATH.exists() and force:
        wb = load_workbook(WORKBOOK_PATH)
    else:
        wb = Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("README")
        for row in README_TEXT:
            ws.append(row)
        ws.column_dimensions["A"].width = 100
        write_sheet(wb, "Aziende", AZIENDE_HEADERS, col_widths=AZIENDE_WIDTHS)
        write_sheet(wb, "Contatti", CONTATTI_HEADERS)
        write_sheet(wb, "Ponti", PONTI_HEADERS)
        write_sheet(wb, "Segnali", SEGNALI_HEADERS)
        write_sheet(wb, "Esiti", ESITI_HEADERS)

    write_sheet(wb, "Config_ICP", ["dimensione", "criterio_operativo", "fonte", "stato"], CONFIG_ICP, [16, 60, 30, 10])
    write_sheet(wb, "Config_AntiICP", ["regola", "motivazione"], CONFIG_ANTI_ICP, [60, 60])
    write_sheet(wb, "Config_Pesi_FIT", ["famiglia", "segnale", "campo_flag_in_Aziende", "punti"], CONFIG_PESI_FIT, [18, 55, 32, 8])
    write_sheet(wb, "Config_Pesi_TIMING", ["famiglia", "segnale", "codice_segnale", "finestra_validita_giorni", "punti", "permanente"], CONFIG_PESI_TIMING, [18, 55, 24, 20, 8, 14])
    write_sheet(wb, "Config_Soglie_Instradamento", ["coda", "condizione", "comportamento", "volume_settimanale_indicativo"], CONFIG_SOGLIE, [8, 40, 55, 22])
    write_sheet(wb, "Config_Ponte_Tipologie", ["tipo_ponte", "come_si_rileva", "forza_0_100"], CONFIG_PONTE_TIPOLOGIE, [65, 40, 12])
    write_sheet(wb, "Config_Segnale_Prodotto", ["segnale_dominante", "area", "frontend", "core_atteso", "ricorsivo_atteso"], CONFIG_SEGNALE_PRODOTTO, [55, 20, 45, 30, 35])

    for name in ["Config_ICP", "Config_AntiICP", "Config_Pesi_FIT", "Config_Pesi_TIMING",
                 "Config_Soglie_Instradamento", "Config_Ponte_Tipologie", "Config_Segnale_Prodotto"]:
        for row in wb[name].iter_rows(min_row=2):
            for cell in row:
                cell.fill = CONFIG_FILL

    order = ["README", "Aziende", "Contatti", "Ponti", "Segnali", "Esiti",
             "Config_ICP", "Config_AntiICP", "Config_Pesi_FIT", "Config_Pesi_TIMING",
             "Config_Soglie_Instradamento", "Config_Ponte_Tipologie", "Config_Segnale_Prodotto"]
    wb._sheets = [wb[n] for n in order]

    wb.save(WORKBOOK_PATH)
    print(f"Workbook creato/aggiornato: {WORKBOOK_PATH}")


if __name__ == "__main__":
    build(force="--force" in sys.argv)
